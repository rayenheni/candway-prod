"""
REST API endpoints for rubric management and deterministic scoring.
"""

import hashlib
import json
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from sqlalchemy import func
from sqlalchemy.orm import Session

from backend.authz import get_application_for_recruiter, get_job_for_recruiter
from backend.database import (
    ABTestAssignment,
    ABTestExperiment,
    Application,
    AuditLog,
    EvaluationResult,
    EvaluationSession,
    Job,
    RubricScoringDetail,
    ScoringVariantResult,
    User,
    get_db,
)
from backend.database import (
    Rubric as RubricDB,
)
from backend.database import (
    SkillDefinition as SkillDefinitionDB,
)
from backend.dependencies import get_current_user, require_admin, require_recruiter
from backend.routers.admin.common import check_permission
from backend.rubric.evidence_analyzer import classify_evidence_quality
from backend.rubric.rubric_engine import score_answer
from backend.rubric.rubric_loader import (
    invalidate_cache,
    load_rubric,
    sync_rubric_skill_definitions,
)
from backend.rubric.rubric_schema import (
    CategoryDefinition,
    JobRubric,
)
from backend.rubric.scoring_aggregator import aggregate_scores
from backend.rubric.skill_mapper import map_extracted_skills
from backend.scoring_service import ScoringService
from backend.tenant import get_current_company_id

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/rubric", tags=["Rubric Engine"])


def _is_admin(user: User) -> bool:
    return bool(user and user.role == "admin")


def _parse_criteria_json(raw) -> list:
    """Parse a Rubric.criteria_json value into a categories list."""
    if not raw:
        return []
    if isinstance(raw, (dict, list)):
        return raw.get("categories", []) if isinstance(raw, dict) else raw
    try:
        data = json.loads(raw)
        return data.get("categories", []) if isinstance(data, dict) else data
    except (json.JSONDecodeError, TypeError):
        return []


def _ensure_job_access(user: User, job_id: int, db: Session) -> Job:
    if not job_id:
        raise HTTPException(
            status_code=400, detail="Application is not linked to a job"
        )
    return get_job_for_recruiter(job_id, user, db)


def _ensure_draft_access(user: User, draft: RubricDB, db: Session) -> None:
    if not _is_admin(user):
        if draft.created_by != user.id:
            raise HTTPException(status_code=403, detail="Not authorized for this draft")
        _ensure_job_access(user, draft.job_id, db)


def _next_draft_version(db: Session, job_id: int) -> int:
    """Return the next negative version for a draft (0, -1, -2, ...)."""
    min_version = (
        db.query(func.min(RubricDB.version))
        .filter(RubricDB.job_id == job_id, RubricDB.version <= 0)
        .scalar()
    )
    return 0 if min_version is None else min_version - 1


def _rubric_json_to_dict(raw: Any) -> dict:
    if isinstance(raw, str):
        return json.loads(raw)
    if isinstance(raw, dict):
        return raw
    return {}


# =============================================================================
# RUBRIC MANAGEMENT
# =============================================================================


@router.post("/jobs/{job_id}")
def create_or_update_rubric(
    job_id: int,
    body: dict,
    request: Request,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Create a new rubric version for a job."""
    job = _ensure_job_access(recruiter, job_id, db)
    # Validate the rubric
    try:
        new_rubric = JobRubric(
            job_id=job_id,
            version=0,  # will be set below
            seniority=body.get("seniority", "mid"),
            categories=[CategoryDefinition(**c) for c in body.get("categories", [])],
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid rubric format: {e}")

    # Get next version number
    latest = (
        db.query(RubricDB)
        .filter(RubricDB.job_id == job_id)
        .order_by(RubricDB.version.desc())
        .first()
    )
    next_version = (latest.version + 1) if latest else 1

    # Mark all previous versions as not active
    db.query(RubricDB).filter(
        RubricDB.job_id == job_id,
        RubricDB.is_active == 1,
    ).update({"is_active": 0})

    new_rubric.version = next_version
    new_rubric.is_current = True

    rubric_company_id = getattr(recruiter, "company_id", None)
    if not rubric_company_id:
        rubric_company_id = getattr(job, "company_id", None)

    db_record = RubricDB(
        job_id=job_id,
        version=next_version,
        is_active=1,
        criteria_json=new_rubric.model_dump_json(),
        created_by=recruiter.id,
        company_id=rubric_company_id,
    )
    db.add(db_record)
    sync_rubric_skill_definitions(
        new_rubric.model_dump_json(), db, company_id=rubric_company_id
    )
    db.commit()
    db.refresh(db_record)

    invalidate_cache(job_id)

    # Link rubric to the already tenant-authorized job
    if job:
        job.rubric_id = db_record.id

    audit = AuditLog(
        user_id=recruiter.id,
        company_id=getattr(recruiter, "_company_id", None),
        action="rubric_create",
        target_id=str(db_record.id),
        details="Created rubric v%d for job %d" % (next_version, job_id),
        ip_address=request.client.host if request.client else None,
    )
    db.add(audit)
    db.commit()

    return {
        "id": db_record.id,
        "job_id": job_id,
        "version": next_version,
        "is_current": True,
        "created_at": db_record.created_at.isoformat(),
    }


@router.get("/jobs/{job_id}")
def get_current_rubric(
    job_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Get the current active rubric for a job."""
    _ensure_job_access(recruiter, job_id, db)
    rubric = load_rubric(job_id)
    if not rubric:
        raise HTTPException(status_code=404, detail="No rubric found for this job")
    return rubric.model_dump()


@router.get("/jobs/{job_id}/versions")
def list_rubric_versions(
    job_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """List all rubric versions for a job."""
    _ensure_job_access(recruiter, job_id, db)
    versions = (
        db.query(RubricDB)
        .filter(RubricDB.job_id == job_id)
        .order_by(RubricDB.version.desc())
        .all()
    )
    return [
        {
            "id": v.id,
            "version": v.version,
            "is_current": bool(v.is_active),
            "created_at": v.created_at.isoformat() if v.created_at else None,
        }
        for v in versions
    ]


@router.get("/jobs/{job_id}/versions/{version}")
def get_rubric_version(
    job_id: int,
    version: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Get a specific rubric version."""
    _ensure_job_access(recruiter, job_id, db)
    record = (
        db.query(RubricDB)
        .filter(RubricDB.job_id == job_id, RubricDB.version == version)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="Rubric version not found")
    import json

    raw = record.criteria_json or ""
    return json.loads(raw) if raw else {}


# =============================================================================
# SCORING
# =============================================================================


@router.post("/interviews/{application_id}/score-answer/{turn_number}")
def score_single_answer(
    application_id: int,
    turn_number: int,
    body: dict,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Score a single answer using the rubric engine."""
    app = get_application_for_recruiter(application_id, recruiter, db)

    eval_session = (
        db.query(EvaluationSession)
        .filter(EvaluationSession.application_id == application_id)
        .first()
    )
    if not eval_session:
        from backend.rubric.interview_starter import InterviewStarter

        eval_session = InterviewStarter.start(db, app)

    from backend.rubric.config_reader import EvaluationConfigReader

    reader = EvaluationConfigReader(eval_session)
    parsed_rubric = reader.get_rubric()
    if not parsed_rubric.raw_json:
        raise HTTPException(
            status_code=400, detail="No rubric configured in the session snapshot"
        )
    from backend.rubric.rubric_schema import JobRubric

    rubric = JobRubric(**parsed_rubric.raw_json)

    extracted_skills = map_extracted_skills(
        body.get("extracted_skills", []),
        rubric.build_lookup(),
    )
    answer_text = body.get("answer_text", "")
    seniority = body.get("seniority", "mid")

    results = score_answer(
        answer_text=answer_text,
        extracted_skills=extracted_skills,
        job_rubric=rubric,
        seniority=seniority,
    )

    # Canonical EvaluationResult creation/upsert.
    # This endpoint only scores an individual answer; final aggregation
    # is performed by score-all through ScoringService.
    eval_result = ScoringService.ensure_score(app, db)

    # Persist scoring details
    for skill_name, result in results.items():
        scoring_row = RubricScoringDetail(
            evaluation_result_id=eval_result.id,
            criterion_name=skill_name,
            criterion_key=result.skill_id,
            score=result.base_score,
            weight=result.quality_multiplier,
            feedback=result.explanation,
        )
        db.add(scoring_row)

    db.commit()

    return {
        "turn_number": turn_number,
        "skills": {k: v.to_dict() for k, v in results.items()},
    }


@router.post("/interviews/{application_id}/score-all")
def score_all_answers(
    application_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Score all answers for an interview and aggregate."""
    app = get_application_for_recruiter(application_id, recruiter, db)

    eval_session = (
        db.query(EvaluationSession)
        .filter(EvaluationSession.application_id == application_id)
        .first()
    )
    if not eval_session:
        raise HTTPException(
            status_code=400, detail="No evaluation session found for this interview"
        )

    from backend.rubric.config_reader import EvaluationConfigReader

    reader = EvaluationConfigReader(eval_session)
    parsed_rubric = reader.get_rubric()
    if not parsed_rubric.raw_json:
        raise HTTPException(
            status_code=400, detail="No rubric configured in the session snapshot"
        )
    from backend.rubric.rubric_schema import JobRubric

    rubric = JobRubric(**parsed_rubric.raw_json)

    existing_results = (
        db.query(RubricScoringDetail)
        .join(
            EvaluationResult,
            RubricScoringDetail.evaluation_result_id == EvaluationResult.id,
        )
        .join(
            EvaluationSession,
            EvaluationResult.evaluation_session_id == EvaluationSession.id,
        )
        .filter(EvaluationSession.application_id == application_id)
        .all()
    )

    if not existing_results:
        raise HTTPException(
            status_code=400, detail="No scoring results found. Submit answers first."
        )

    # RubricScoringDetail does not currently persist a turn_id.
    # Therefore we cannot reconstruct the original per-turn grouping here.
    # Aggregate the persisted criterion scores as one canonical scoring set
    # instead of incorrectly treating every detail-row PK as a separate turn.
    all_scores: dict = {0: {}}

    from backend.rubric.rubric_engine import SkillScoreResult

    for row in existing_results:
        result = SkillScoreResult(
            skill_name=row.criterion_name,
            skill_id=row.criterion_key or "",
            base_score=float(row.score),
            quality="medium",
            quality_multiplier=1.0,
            final_score=float(row.score),
            confidence_lower=0,
            confidence_upper=0,
            evidence_sentences=[],
            matched_level="",
            matched_keywords=[],
            missing_competencies=[],
            explanation=row.feedback or "",
        )

        # Keep the strongest/latest persisted score for duplicate criteria.
        current = all_scores[0].get(row.criterion_name)
        if current is None or float(row.score) >= float(current.final_score):
            all_scores[0][row.criterion_name] = result

    seniority = "mid"
    summary = aggregate_scores(
        application_id=application_id,
        rubric=rubric,
        all_answer_results=all_scores,
        seniority=seniority,
    )

    break_down = {
        "category_scores": [c.to_dict() for c in summary.categories],
        "skill_scores": summary.skill_scores,
        "gaps": summary.gaps,
        "num_answers_scored": summary.num_answers_scored,
    }

    ScoringService.compute_final_score(
        app,
        db,
        computed_by="rubric_router",
        override_rubric_score=float(summary.overall_score),
        override_rubric_coverage_pct=float(summary.overall_coverage_pct),
        extra_breakdown=break_down,
        confidence_lower=summary.confidence_lower,
        confidence_upper=summary.confidence_upper,
    )

    db.commit()

    return summary.to_dict()


@router.get("/interviews/{application_id}/summary")
def get_scoring_summary(
    application_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Get the full scoring summary for an interview."""
    _app = get_application_for_recruiter(application_id, recruiter, db)
    summary = (
        db.query(EvaluationResult)
        .join(
            EvaluationSession,
            EvaluationResult.evaluation_session_id == EvaluationSession.id,
        )
        .filter(EvaluationSession.application_id == application_id)
        .first()
    )
    if not summary:
        raise HTTPException(
            status_code=404,
            detail="No scoring summary found. Score the interview first.",
        )

    breakdown = summary.score_breakdown or {}
    if isinstance(breakdown, str):
        import json

        breakdown = json.loads(breakdown) if breakdown else {}
    return {
        "application_id": application_id,
        "rubric_version": summary.rubric_version,
        "overall_score": summary.final_score,
        "confidence_range": [summary.confidence_lower, summary.confidence_upper],
        "categories": breakdown.get("category_scores", [])
        if isinstance(breakdown, dict)
        else [],
        "skill_scores": breakdown.get("skill_scores", {})
        if isinstance(breakdown, dict)
        else {},
        "gaps": breakdown.get("gaps", []) if isinstance(breakdown, dict) else [],
        "num_answers_scored": breakdown.get("num_answers_scored", 0)
        if isinstance(breakdown, dict)
        else 0,
        "computed_at": summary.computed_at.isoformat() if summary.computed_at else None,
    }


@router.get("/interviews/{application_id}/explain/{skill_name}")
def get_skill_explanation(
    application_id: int,
    skill_name: str,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Get explanation for a specific skill across all answers."""
    _app = get_application_for_recruiter(application_id, recruiter, db)
    results = (
        db.query(RubricScoringDetail)
        .join(
            EvaluationResult,
            RubricScoringDetail.evaluation_result_id == EvaluationResult.id,
        )
        .join(
            EvaluationSession,
            EvaluationResult.evaluation_session_id == EvaluationSession.id,
        )
        .filter(
            EvaluationSession.application_id == application_id,
            RubricScoringDetail.criterion_name == skill_name.lower(),
        )
        .order_by(RubricScoringDetail.id)
        .all()
    )

    if not results:
        raise HTTPException(
            status_code=404, detail=f"No results for skill '{skill_name}'"
        )

    answers = []
    for r in results:
        answers.append(
            {
                "answer_id": r.id,
                "final_score": r.score,
                "evidence": [],
                "matched_level": "",
                "missing_competencies": [],
                "explanation": r.feedback,
            }
        )

    return {
        "skill_name": skill_name,
        "answers": answers,
        "best_score": max(a["final_score"] for a in answers) if answers else 0,
        "improvement_tip": _generate_improvement_tip(skill_name, answers),
    }


def _generate_improvement_tip(skill_name: str, answers: list) -> str:
    missing_all = set()
    for a in answers:
        for m in a.get("missing_competencies", []):
            missing_all.add(m)

    if missing_all:
        return f"Focus on: {', '.join(list(missing_all)[:3])}"
    return "Review the level descriptors for this skill."


@router.get("/jobs/{job_id}/compare")
def compare_candidates(
    job_id: int,
    candidate_ids: str = Query(..., description="Comma-separated application IDs"),
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Compare multiple candidates for the same job."""
    _ensure_job_access(recruiter, job_id, db)
    ids = [int(x.strip()) for x in candidate_ids.split(",") if x.strip()]
    if not ids:
        raise HTTPException(status_code=400, detail="No candidate IDs provided")

    summaries = (
        db.query(EvaluationResult)
        .join(
            EvaluationSession,
            EvaluationResult.evaluation_session_id == EvaluationSession.id,
        )
        .filter(EvaluationSession.application_id.in_(ids))
        .all()
    )

    summary_map = {s.evaluation_session.application_id: s for s in summaries}

    candidates = []
    for app_id in ids:
        summary = summary_map.get(app_id)
        try:
            app = get_application_for_recruiter(app_id, recruiter, db)
        except HTTPException:
            continue

        if not summary or not app:
            continue
        if app.job_id != job_id:
            continue

        breakdown = summary.score_breakdown or {}
        if isinstance(breakdown, str):
            import json

            breakdown = json.loads(breakdown) if breakdown else {}
        bd = breakdown if isinstance(breakdown, dict) else {}

        cat_scores = {}
        cats = bd.get("category_scores", [])
        if cats:
            for c in cats:
                if isinstance(c, dict):
                    cat_scores[c.get("name", "Unknown")] = c.get("score", 0)

        all_skills = {}
        skill_scores = bd.get("skill_scores", {})
        if skill_scores:
            all_skills = skill_scores

        sorted_skills = sorted(
            all_skills.items(),
            key=lambda x: x[1].get("final_score", 0) if isinstance(x[1], dict) else 0,
            reverse=True,
        )
        top_skills = [
            f"{s[0]} ({s[1].get('final_score', s[1]) if isinstance(s[1], dict) else s[1]})"
            for s in sorted_skills[:3]
        ]
        weak_skills = (
            [
                f"{s[0]} ({s[1].get('final_score', s[1]) if isinstance(s[1], dict) else s[1]})"
                for s in sorted_skills[-3:]
            ]
            if len(sorted_skills) >= 3
            else []
        )

        candidates.append(
            {
                "id": app.id,
                "name": app.full_name or f"Candidate #{app.id}",
                "overall_score": summary.final_score,
                "confidence_range": [
                    summary.confidence_lower,
                    summary.confidence_upper,
                ],
                "categories": cat_scores,
                "top_skills": top_skills,
                "weakest_skills": weak_skills,
                "gaps": bd.get("gaps", []),
            }
        )

    rubric = load_rubric(job_id)

    return {
        "job_id": job_id,
        "rubric_version": rubric.version if rubric else None,
        "candidates": sorted(
            candidates, key=lambda c: c["overall_score"], reverse=True
        ),
    }


@router.get("/taxonomy")
def get_taxonomy():
    """Get the current default taxonomy as a reference."""
    try:
        rubric = load_rubric(0)
        if rubric:
            return rubric.model_dump()
    except Exception as e:
        logger.error(f"Taxonomy load failed: {e}")
    return {"categories": []}


@router.post("/import")
def import_rubric_excel(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Import a rubric from an Excel template file."""
    check_permission(current_user, "manage_content")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file.file)
        ws = wb.active

        # Expected columns: Categorie, Sous-categories, Roles, Criteres evaluation, Competences, Methodes entretien
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        categories_map = {}  # category_name -> CategoryDefinition dict
        subcat_map = {}  # (category_name, subcat_name) -> subcategory dict

        for row in rows:
            if not row or not row[0]:
                continue
            cat_name = str(row[0]).strip() if row[0] else ""
            subcat_name = str(row[1]).strip() if row[1] else ""
            roles_str = str(row[2]).strip() if row[2] else ""
            criteria_str = str(row[3]).strip() if row[3] else ""
            skills_str = str(row[4]).strip() if row[4] else ""
            methods_str = str(row[5]).strip() if row[5] else ""

            if not cat_name:
                continue

            # Initialize category if new
            if cat_name not in categories_map:
                categories_map[cat_name] = {
                    "name": cat_name,
                    "description": "",
                    "weight": 1.0,
                    "evaluation_criteria": [],
                    "interview_methods": [],
                    "target_roles": [],
                    "subcategories": [],
                }

            cat = categories_map[cat_name]

            # Collect evaluation criteria (comma-separated)
            if criteria_str:
                for c in criteria_str.split(","):
                    c = c.strip()
                    if c and c not in cat["evaluation_criteria"]:
                        cat["evaluation_criteria"].append(c)

            # Collect interview methods (comma-separated)
            if methods_str:
                for m in methods_str.split(","):
                    m = m.strip()
                    if m and m not in cat["interview_methods"]:
                        cat["interview_methods"].append(m)

            # Collect target roles (comma-separated)
            if roles_str:
                for r in roles_str.split(","):
                    r = r.strip()
                    if r and r not in cat["target_roles"]:
                        cat["target_roles"].append(r)

            if subcat_name:
                subcat_key = (cat_name, subcat_name)
                if subcat_key not in subcat_map:
                    subcat_map[subcat_key] = {
                        "name": subcat_name,
                        "description": "",
                        "weight": 1.0,
                        "skills": [],
                    }

                # Add skills to subcategory
                if skills_str:
                    for s in skills_str.split(","):
                        s = s.strip()
                        if s:
                            skill_names = [
                                sk["name"].lower()
                                for sk in subcat_map[subcat_key]["skills"]
                            ]
                            if s.lower() not in skill_names:
                                subcat_map[subcat_key]["skills"].append(
                                    {
                                        "name": s,
                                        "description": "",
                                        "weight": 1.0,
                                        "is_required": False,
                                        "keywords": [s.lower()],
                                        "levels": {
                                            "junior": [
                                                {
                                                    "score_threshold": 30,
                                                    "description": f"Basic {s}",
                                                    "keywords": [f"basic {s.lower()}"],
                                                    "sort_order": 1,
                                                },
                                                {
                                                    "score_threshold": 60,
                                                    "description": f"Working {s}",
                                                    "keywords": [
                                                        f"{s.lower()} experience"
                                                    ],
                                                    "sort_order": 2,
                                                },
                                                {
                                                    "score_threshold": 90,
                                                    "description": f"Expert {s}",
                                                    "keywords": [f"expert {s.lower()}"],
                                                    "sort_order": 3,
                                                },
                                            ],
                                            "mid": [
                                                {
                                                    "score_threshold": 30,
                                                    "description": f"Basic {s}",
                                                    "keywords": [f"basic {s.lower()}"],
                                                    "sort_order": 1,
                                                },
                                                {
                                                    "score_threshold": 60,
                                                    "description": f"Working {s}",
                                                    "keywords": [
                                                        f"{s.lower()} experience"
                                                    ],
                                                    "sort_order": 2,
                                                },
                                                {
                                                    "score_threshold": 90,
                                                    "description": f"Expert {s}",
                                                    "keywords": [f"expert {s.lower()}"],
                                                    "sort_order": 3,
                                                },
                                            ],
                                            "senior": [
                                                {
                                                    "score_threshold": 30,
                                                    "description": f"Basic {s}",
                                                    "keywords": [f"basic {s.lower()}"],
                                                    "sort_order": 1,
                                                },
                                                {
                                                    "score_threshold": 60,
                                                    "description": f"Working {s}",
                                                    "keywords": [
                                                        f"{s.lower()} experience"
                                                    ],
                                                    "sort_order": 2,
                                                },
                                                {
                                                    "score_threshold": 90,
                                                    "description": f"Expert {s}",
                                                    "keywords": [f"expert {s.lower()}"],
                                                    "sort_order": 3,
                                                },
                                            ],
                                        },
                                    }
                                )

        # Build category list with subcategories
        categories = []
        for cat_name, cat in categories_map.items():
            cat["subcategories"] = [
                v for k, v in subcat_map.items() if k[0] == cat_name
            ]
            categories.append(cat)

        # Create draft
        draft = RubricDB(
            job_id=0,  # 0 = standalone template
            version=0,
            is_active=0,
            created_by=current_user.id,
            title=file.filename.replace(".xlsx", "").replace(".xls", ""),
            criteria_json={
                "version": 1,
                "seniority": "mid",
                "categories": categories,
            },
        )
        db.add(draft)
        db.commit()
        db.refresh(draft)

        audit = AuditLog(
            user_id=current_user.id,
            company_id=getattr(current_user, "_company_id", None),
            action="rubric_import",
            target_id=str(draft.id),
            details="Imported rubric from file '%s' (%d categories, %d skills)"
            % (
                file.filename,
                len(categories),
                sum(len(sk["skills"]) for sk in subcat_map.values()),
            ),
            ip_address=request.client.host if request.client else None,
        )
        db.add(audit)
        db.commit()

        return {
            "success": True,
            "draft_id": draft.id,
            "title": draft.title,
            "categories_count": len(categories),
            "skills_count": sum(len(sk["skills"]) for sk in subcat_map.values()),
        }

    except Exception as e:
        logger.error(f"[RUBRIC-IMPORT] Failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=400, detail=f"Failed to parse Excel file: {str(e)}"
        )


@router.get("/export/{job_id}")
def export_rubric_excel(
    job_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export a rubric as an Excel file."""
    import io

    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    check_permission(current_user, "manage_content")

    rubric_record = (
        db.query(RubricDB)
        .filter(RubricDB.job_id == job_id, RubricDB.is_active == 1)
        .first()
    )
    if rubric_record:
        raw = rubric_record.criteria_json or ""
        criteria_dict = (
            json.loads(raw)
            if isinstance(raw, str) and raw
            else (raw if isinstance(raw, dict) else {})
        )
        rubric_data = {
            "job_id": rubric_record.job_id,
            "version": rubric_record.version,
            "categories": criteria_dict.get(
                "criteria",
                criteria_dict.get("categories", criteria_dict.get("sections", [])),
            ),
            "seniority": rubric_record.complexity or "mid",
        }
    else:
        # Try loading from draft (is_active=0 indicates draft)
        rubric_data = None
        draft = (
            db.query(RubricDB)
            .filter(
                RubricDB.job_id == job_id,
                RubricDB.is_active == 0,
            )
            .order_by(RubricDB.updated_at.desc())
            .first()
        )
        if draft:
            rubric_data = draft.criteria_json
        else:
            raise HTTPException(status_code=404, detail="No rubric found for this job")

    rubric = JobRubric(**rubric_data) if isinstance(rubric_data, dict) else rubric_data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rubric Export"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="6366F1", end_color="6366F1", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "Catégorie",
        "Sous-catégories",
        "Rôles",
        "Critères d'évaluation",
        "Compétences",
        "Méthodes d'entretien",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    widths = [25, 25, 30, 35, 35, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    row = 2
    for cat in rubric.categories:
        criteria_str = (
            ", ".join(cat.evaluation_criteria)
            if hasattr(cat, "evaluation_criteria") and cat.evaluation_criteria
            else ""
        )
        methods_str = (
            ", ".join(cat.interview_methods)
            if hasattr(cat, "interview_methods") and cat.interview_methods
            else ""
        )
        roles_str = (
            ", ".join(cat.target_roles)
            if hasattr(cat, "target_roles") and cat.target_roles
            else ""
        )

        if cat.subcategories:
            for sub in cat.subcategories:
                skills_str = ", ".join(s.name for s in sub.skills) if sub.skills else ""
                ws.cell(row=row, column=1, value=cat.name).border = thin_border
                ws.cell(row=row, column=2, value=sub.name).border = thin_border
                ws.cell(row=row, column=3, value=roles_str).border = thin_border
                ws.cell(row=row, column=4, value=criteria_str).border = thin_border
                ws.cell(row=row, column=5, value=skills_str).border = thin_border
                ws.cell(row=row, column=6, value=methods_str).border = thin_border
                for col in range(1, 7):
                    ws.cell(row=row, column=col).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
                row += 1
        else:
            ws.cell(row=row, column=1, value=cat.name).border = thin_border
            ws.cell(row=row, column=3, value=roles_str).border = thin_border
            ws.cell(row=row, column=4, value=criteria_str).border = thin_border
            ws.cell(row=row, column=6, value=methods_str).border = thin_border
            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            row += 1

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=rubric_job_{job_id}.xlsx"
        },
    )


# =============================================================================
# PHASE 4 — MANAGEMENT (table view for admin/recruiter)
# =============================================================================


def _count_skills(categories: list) -> int:
    if not categories:
        return 0
    total = 0
    for cat in categories:
        for sub in cat.get("subcategories", []) or [cat]:
            total += len(sub.get("skills", []) or [])
    return total


@router.get("/management")
def get_management_view(
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Return jobs + their rubrics + drafts in one payload for the management table."""
    # Jobs the user can see — tenant scoped for non-admins.
    job_q = db.query(Job)
    if not _is_admin(recruiter):
        company_id = getattr(recruiter, "_company_id", None)
        if company_id is None:
            from backend.authz import _user_company_id

            company_id = _user_company_id(db, recruiter.id)

        if company_id is None:
            raise HTTPException(status_code=404, detail="No company membership")

        job_q = job_q.filter(Job.company_id == company_id)

    jobs = job_q.order_by(Job.created_at.desc()).all()
    job_ids = [j.id for j in jobs]

    # Latest rubric per job
    rubrics_by_job: dict = {}
    if job_ids:
        rubric_rows = (
            db.query(RubricDB)
            .filter(RubricDB.job_id.in_(job_ids))
            .order_by(RubricDB.job_id, RubricDB.version.desc())
            .all()
        )
        for r in rubric_rows:
            if r.job_id not in rubrics_by_job:
                raw = r.criteria_json or ""
                if isinstance(raw, str):
                    try:
                        import json as _json

                        raw = _json.loads(raw) if raw else {}
                    except Exception:
                        raw = {}
                cats = (
                    (raw or {}).get(
                        "criteria", raw.get("categories", raw.get("sections", []))
                    )
                    if isinstance(raw, dict)
                    else []
                )
                rubrics_by_job[r.job_id] = {
                    "id": r.id,
                    "version": r.version,
                    "seniority": r.complexity or "mid",
                    "category_count": len(cats),
                    "skill_count": _count_skills(cats),
                    "is_current": bool(r.is_active),
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                }

    # Drafts for this user (is_active=0 indicates draft)
    draft_q = db.query(RubricDB).filter(
        RubricDB.is_active == 0,
        RubricDB.created_by == recruiter.id,
    )
    drafts = draft_q.order_by(RubricDB.updated_at.desc()).limit(50).all()
    drafts_out = []
    for d in drafts:
        drafts_out.append(
            {
                "id": d.id,
                "title": d.title,
                "job_id": d.job_id,
                "updated_at": d.updated_at.isoformat() if d.updated_at else None,
            }
        )

    # Application counts per job
    app_counts: dict = {}
    if job_ids:
        from sqlalchemy import func

        rows = (
            db.query(Application.job_id, func.count(Application.id))
            .filter(Application.job_id.in_(job_ids))
            .group_by(Application.job_id)
            .all()
        )
        for jid, cnt in rows:
            app_counts[jid] = cnt

    rows_out = []
    for j in jobs:
        rubric = rubrics_by_job.get(j.id)
        rows_out.append(
            {
                "job_id": j.id,
                "job_title": j.title,
                "is_active": bool(j.is_active),
                "location": j.location,
                "type": j.type,
                "application_count": app_counts.get(j.id, 0),
                "rubric": rubric,  # null = no rubric yet
                "has_draft": any(d["job_id"] == j.id for d in drafts_out),
            }
        )

    return {
        "rows": rows_out,
        "drafts": drafts_out,
        "stats": {
            "total_jobs": len(rows_out),
            "with_rubric": sum(1 for r in rows_out if r["rubric"]),
            "without_rubric": sum(1 for r in rows_out if not r["rubric"]),
            "drafts": len(drafts_out),
            "total_skills": sum(
                (r["rubric"] or {}).get("skill_count", 0) for r in rows_out
            ),
        },
    }


@router.get("/templates")
def list_rubric_templates(
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
    company_id: int = Depends(get_current_company_id),
):
    """List published rubrics for jobs owned by this recruiter's company.

    Covers both rubric storage modes:
    - legacy job-bound rubrics (`Rubric.job_id` set)
    - standalone library rubrics (`Rubric.job_id IS NULL`) linked to a job
      via `Job.rubric_id`
    """
    owned_jobs = (
        db.query(Job)
        .filter(
            Job.recruiter_id == recruiter.id,
            Job.company_id == company_id,
        )
        .all()
    )
    owned_job_ids = [j.id for j in owned_jobs]

    # Jobs grouped by their linked standalone rubric id.
    job_by_linked_rubric = {}
    for j in owned_jobs:
        if j.rubric_id:
            job_by_linked_rubric.setdefault(j.rubric_id, j)

    # Rubrics directly bound to owned jobs (legacy flow, RubricDB.job_id set).
    bound_rubrics = []
    if owned_job_ids:
        bound_rubrics = (
            db.query(RubricDB)
            .filter(
                RubricDB.job_id.in_(owned_job_ids),
                RubricDB.is_active == 1,
            )
            .order_by(RubricDB.job_id, RubricDB.version.desc())
            .all()
        )

    # Standalone library rubrics linked via Job.rubric_id (new flow).
    linked_rubrics = []
    linked_rubric_ids = list(job_by_linked_rubric.keys())
    if linked_rubric_ids:
        linked_rubrics = (
            db.query(RubricDB)
            .filter(
                RubricDB.id.in_(linked_rubric_ids),
                RubricDB.is_active == 1,
            )
            .all()
        )

    # For bound rubrics, the job is RubricDB.job_id; for linked rubrics the
    # job is the one whose rubric_id references this row. Job-bound entries win
    # when a rubric is both bound and linked (dedup by rubric id).
    job_id_by_rubric_id = {}
    job_by_id = {j.id: j for j in owned_jobs}
    for r in linked_rubrics:
        job_id_by_rubric_id[r.id] = job_by_linked_rubric[r.id].id
    for r in bound_rubrics:
        job_id_by_rubric_id[r.id] = r.job_id

    results = []
    seen_rubric_ids = set()
    for rubric in bound_rubrics + linked_rubrics:
        if rubric.id in seen_rubric_ids:
            continue
        seen_rubric_ids.add(rubric.id)
        job = job_by_id.get(job_id_by_rubric_id.get(rubric.id))
        if not job:
            continue
        cats = _parse_criteria_json(rubric.criteria_json)
        total_skills = sum(
            len(sc.get("skills", []))
            for cat in cats
            for sc in cat.get("subcategories", [])
        )
        results.append(
            {
                "job_id": job.id,
                "job_title": rubric.title or job.title or f"Job #{job.id}",
                "company": job.company_name,
                "rubric_id": rubric.id,
                "version": rubric.version,
                "seniority": rubric.complexity or "mid",
                "category_count": len(cats),
                "skill_count": total_skills,
                "categories": cats,
            }
        )

    results.sort(key=lambda r: (r["job_title"] or "").lower())
    return {"templates": results}


@router.get("/template-detail/{job_id}")
def get_template_detail(
    job_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Return the full rubric data for a job owned by this recruiter."""
    get_job_for_recruiter(job_id, recruiter, db)
    rubric = load_rubric(job_id)
    if not rubric:
        raise HTTPException(status_code=404, detail="No rubric found for this job")
    return rubric.model_dump()


@router.post("/duplicate/{job_id}")
def duplicate_rubric(
    job_id: int,
    request: Request,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Clone the current published rubric for a job into a new editable draft."""
    job = _ensure_job_access(recruiter, job_id, db)
    current = load_rubric(job_id)
    if not current:
        raise HTTPException(status_code=404, detail="No published rubric to duplicate")
    next_ver = _next_draft_version(db, job_id)
    draft = RubricDB(
        job_id=job_id,
        version=next_ver,
        is_active=0,
        created_by=recruiter.id,
        title=f"{job.title} (copy)",
        criteria_json=current.model_dump(),
        company_id=job.company_id,
    )
    db.add(draft)
    db.commit()
    db.refresh(draft)

    audit = AuditLog(
        user_id=recruiter.id,
        company_id=getattr(recruiter, "_company_id", None),
        action="rubric_duplicate",
        target_id=str(draft.id),
        details="Duplicated rubric for job %d as draft v%d" % (job_id, next_ver),
        ip_address=request.client.host if request.client else None,
    )
    db.add(audit)
    db.commit()

    return {"id": draft.id, "title": draft.title, "job_id": job_id}


# =============================================================================
# PHASE 4 — GUIDED RUBRIC GENERATION (AI-ASSISTED)
# =============================================================================


_GENERATE_PROMPT = """Task: Build a concise skill rubric for the role below. Output a JSON rubric skeleton only.

Role context:
{context}

Return strictly a JSON object with this exact shape (no extra keys, no commentary):
{{
  "role_title": "<short role title>",
  "seniority": "junior" | "mid" | "senior",
  "categories": [
    {{
      "name": "<Category Name>",
      "weight": <int 0-100, weights across categories sum to 100>,
      "skills": [
        {{
          "name": "<Skill Name>",
          "weight": <int 0-100, weights across skills within a category sum to the category weight>,
          "is_required": <bool>,
          "keywords": ["<kw1>", "<kw2>", "<kw3>"]
        }}
      ]
    }}
  ],
  "suggested_extra_skills": ["<skill>", "<skill>", "<skill>"]
}}

Constraints:
- 3 to 5 categories
- 2 to 4 skills per category
- Category weights sum to 100
- Within each category, skill weights sum to the category weight
- All keywords lowercase, single words or short phrases
- Be specific to the role (e.g. "React" not "Frontend", "PostgreSQL" not "Database")
"""


def _fallback_rubric(role_title: str) -> dict:
    """Deterministic skeleton used when the LLM is unavailable."""
    title = (role_title or "General Role").strip() or "General Role"
    return {
        "role_title": title,
        "seniority": "mid",
        "categories": [
            {
                "name": "Core Skills",
                "weight": 50,
                "skills": [
                    {
                        "name": "Problem Solving",
                        "weight": 25,
                        "is_required": True,
                        "keywords": ["problem", "solve", "analyze", "debug"],
                    },
                    {
                        "name": "Communication",
                        "weight": 25,
                        "is_required": True,
                        "keywords": [
                            "communicate",
                            "explain",
                            "collaborate",
                            "present",
                        ],
                    },
                ],
            },
            {
                "name": "Domain Knowledge",
                "weight": 30,
                "skills": [
                    {
                        "name": f"{title} Fundamentals",
                        "weight": 30,
                        "is_required": True,
                        "keywords": [title.lower(), "fundamentals", "core"],
                    },
                ],
            },
            {
                "name": "Tools & Workflow",
                "weight": 20,
                "skills": [
                    {
                        "name": "Version Control",
                        "weight": 10,
                        "is_required": False,
                        "keywords": ["git", "branch", "merge", "commit"],
                    },
                    {
                        "name": "Testing",
                        "weight": 10,
                        "is_required": False,
                        "keywords": ["test", "verify", "validate", "qa"],
                    },
                ],
            },
        ],
        "suggested_extra_skills": ["Documentation", "Code Review", "Mentoring"],
    }


@router.post("/generate")
async def generate_rubric(
    body: dict,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """AI-generate a rubric skeleton from a job, JD, or role title.

    Body shape (exactly one of these should be provided):
      { "job_id": <int> }   — use an existing job's title/description
      { "jd_text": "..." }  — use a pasted job description
      { "role_title": "..." } — describe a role freely
    """
    job_id = body.get("job_id")
    jd_text = (body.get("jd_text") or "").strip()
    role_title = (body.get("role_title") or "").strip()

    context_parts = []
    if job_id:
        job = _ensure_job_access(recruiter, job_id, db)
        role_title = job.title or role_title
        if job.description:
            context_parts.append(f"Job description:\n{job.description[:2000]}")
    elif jd_text:
        context_parts.append(f"Job description:\n{jd_text[:2000]}")
    elif role_title:
        context_parts.append(f"Role title: {role_title}")
    else:
        raise HTTPException(
            status_code=400,
            detail="Provide one of: job_id, jd_text, or role_title",
        )

    context_parts.append(f"Role title: {role_title or 'Unknown'}")
    context = "\n\n".join(context_parts)

    try:
        from backend.ai.llm import call_groq_cascade

        raw = await call_groq_cascade(
            [
                {"role": "system", "content": "Output: rubric skeleton as JSON only."},
                {"role": "user", "content": _GENERATE_PROMPT.format(context=context)},
            ],
            json_mode=True,
        )
        if isinstance(raw, str):
            import json as _json

            data = _json.loads(raw)
        else:
            data = raw
        if not isinstance(data, dict) or "categories" not in data:
            raise ValueError("LLM returned unexpected shape")
        data.setdefault("role_title", role_title or "Role")
        data.setdefault("seniority", "mid")
        data.setdefault("suggested_extra_skills", [])
        data["_source"] = "llm"
    except Exception as e:
        logger.warning(f"[generate_rubric] LLM failed, using fallback: {e}")
        data = _fallback_rubric(role_title or "Role")
        data["_source"] = "fallback"

    return data


# =============================================================================
# PHASE 2 — DRAFT MANAGEMENT
# =============================================================================


@router.post("/drafts/{job_id}")
def create_draft(
    job_id: int,
    body: dict,
    request: Request,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Create a new rubric draft for a job."""
    job = _ensure_job_access(recruiter, job_id, db)
    current = load_rubric(job_id)
    rubric_json = (
        current.model_dump()
        if current
        else JobRubric(job_id=job_id, version=1, categories=[]).model_dump()
    )

    next_ver = _next_draft_version(db, job_id)
    draft = RubricDB(
        job_id=job_id,
        version=next_ver,
        is_active=0,
        created_by=recruiter.id,
        title=body.get("name", "Untitled Draft"),
        criteria_json=rubric_json,
        company_id=job.company_id,
    )
    db.add(draft)
    db.commit()
    db.refresh(draft)

    audit = AuditLog(
        user_id=recruiter.id,
        company_id=getattr(recruiter, "_company_id", None),
        action="rubric_draft_create",
        target_id=str(draft.id),
        details="Created rubric draft for job %d (v%d)" % (job_id, next_ver),
        ip_address=request.client.host if request.client else None,
    )
    db.add(audit)
    db.commit()

    return {
        "id": draft.id,
        "title": draft.title,
        "created_at": draft.created_at.isoformat(),
    }


@router.put("/drafts/{draft_id}")
def save_draft(
    draft_id: int,
    body: dict,
    request: Request,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Save a rubric draft."""
    draft = (
        db.query(RubricDB)
        .filter(
            RubricDB.id == draft_id,
            RubricDB.is_active == 0,
        )
        .first()
    )
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    _ensure_draft_access(recruiter, draft, db)
    draft.criteria_json = body.get("rubric_json", draft.criteria_json)
    draft.title = body.get("name", draft.title)
    db.commit()

    audit = AuditLog(
        user_id=recruiter.id,
        company_id=getattr(recruiter, "_company_id", None),
        action="rubric_draft_update",
        target_id=str(draft_id),
        details="Updated rubric draft %d for job %d" % (draft_id, draft.job_id),
        ip_address=request.client.host if request.client else None,
    )
    db.add(audit)
    db.commit()

    return {
        "id": draft.id,
        "title": draft.title,
        "updated_at": datetime.utcnow().isoformat(),
    }


@router.get("/drafts/{draft_id}")
def get_draft(
    draft_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Get a rubric draft."""
    draft = (
        db.query(RubricDB)
        .filter(
            RubricDB.id == draft_id,
            RubricDB.is_active == 0,
        )
        .first()
    )
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    _ensure_draft_access(recruiter, draft, db)
    return {
        "id": draft.id,
        "job_id": draft.job_id,
        "title": draft.title,
        "criteria_json": draft.criteria_json,
        "updated_at": draft.updated_at.isoformat() if draft.updated_at else None,
    }


@router.get("/drafts")
def list_drafts(
    job_id: Optional[int] = Query(None),
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """List drafts for a recruiter."""
    q = db.query(RubricDB).filter(
        RubricDB.is_active == 0,
        RubricDB.created_by == recruiter.id,
    )
    if job_id:
        q = q.filter(RubricDB.job_id == job_id)
    drafts = q.order_by(RubricDB.updated_at.desc()).all()
    return [
        {
            "id": d.id,
            "job_id": d.job_id,
            "title": d.title,
            "updated_at": d.updated_at.isoformat() if d.updated_at else None,
        }
        for d in drafts
    ]


@router.post("/drafts/{draft_id}/publish")
def publish_draft(
    draft_id: int,
    body: dict,
    request: Request,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Publish a draft as a new rubric version."""
    draft = (
        db.query(RubricDB)
        .filter(
            RubricDB.id == draft_id,
            RubricDB.is_active == 0,
        )
        .first()
    )
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    _ensure_draft_access(recruiter, draft, db)

    job_id = draft.job_id
    latest = (
        db.query(RubricDB)
        .filter(RubricDB.job_id == job_id)
        .order_by(RubricDB.version.desc())
        .first()
    )
    next_version = (latest.version + 1) if latest else 1

    try:
        draft_json = _rubric_json_to_dict(draft.criteria_json)
        rubric = JobRubric(
            job_id=job_id,
            version=next_version,
            seniority=body.get("seniority", "mid"),
            categories=[
                CategoryDefinition(**c) for c in draft_json.get("categories", [])
            ],
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid rubric: {e}")

    db.query(RubricDB).filter(
        RubricDB.job_id == job_id,
        RubricDB.is_active == 1,
    ).update({"is_active": 0})

    rubric_company_id = getattr(draft, "company_id", None) or getattr(
        recruiter, "company_id", None
    )

    rubric_record = RubricDB(
        job_id=job_id,
        version=next_version,
        is_active=1,
        criteria_json=rubric.model_dump_json(),
        created_by=recruiter.id,
        company_id=rubric_company_id,
    )
    db.add(rubric_record)
    sync_rubric_skill_definitions(
        rubric.model_dump_json(), db, company_id=rubric_company_id
    )
    db.commit()

    invalidate_cache(job_id)

    audit = AuditLog(
        user_id=recruiter.id,
        company_id=getattr(recruiter, "_company_id", None),
        action="rubric_publish",
        target_id=str(rubric_record.id),
        details="Published rubric v%d for job %d (from draft %d)"
        % (next_version, job_id, draft_id),
        ip_address=request.client.host if request.client else None,
    )
    db.add(audit)
    db.commit()

    return {
        "rubric_id": rubric_record.id,
        "version": next_version,
        "draft_id": draft_id,
    }


# =============================================================================
# CATEGORY → JOB LOOKUP
# =============================================================================


@router.get("/job-for-category/{category_id}")
def job_for_category(
    category_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Find a job linked to the given category. Requires recruiter authentication with company ownership check."""
    job = (
        db.query(Job)
        .filter(Job.category_id == category_id, Job.deleted_at.is_(None))
        .first()
    )
    if not job:
        raise HTTPException(status_code=404, detail="No job found for this category")
    _ensure_job_access(recruiter, job.id, db)
    return {"job_id": job.id, "title": job.title}


# =============================================================================
# PHASE 2 — LIVE SCORING PREVIEW
# =============================================================================


@router.post("/preview-score")
def preview_score(
    body: dict,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Preview rubric score for a raw answer without persisting.
    Supports optional variant scoring for A/B comparison.
    """
    job_id = body.get("job_id")
    answer_text = body.get("answer_text", "")
    seniority = body.get("seniority", "mid")
    extracted_skills_raw = body.get("extracted_skills", None)
    variant_rubric = body.get("variant_rubric", None)

    if not job_id:
        raise HTTPException(status_code=400, detail="job_id is required")
    if not answer_text.strip():
        raise HTTPException(status_code=400, detail="answer_text is required")
    _ensure_job_access(recruiter, job_id, db)

    rubric = load_rubric(job_id)
    if not rubric:
        raise HTTPException(status_code=400, detail="No rubric configured for this job")

    if extracted_skills_raw:
        extracted = extracted_skills_raw
    else:
        extracted = _mock_extract(answer_text)

    mapped = map_extracted_skills(extracted, rubric.build_lookup())
    for item in mapped:
        quality, reason = classify_evidence_quality(
            item.get("evidence_sentences", []),
            item.get("skill_name", ""),
        )
        item["quality"] = quality
        item["quality_reason"] = reason

    scores_a = score_answer(
        answer_text=answer_text,
        extracted_skills=mapped,
        job_rubric=rubric,
        seniority=seniority,
    )

    scores_b = None
    if variant_rubric:
        try:
            vrub = JobRubric(**variant_rubric)
            scores_b = score_answer(
                answer_text=answer_text,
                extracted_skills=mapped,
                job_rubric=vrub,
                seniority=seniority,
            )
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid variant rubric: {e}")

    score_list_a = [s.final_score for s in scores_a.values()]
    overall_a = int(sum(score_list_a) / len(score_list_a)) if score_list_a else 0

    result = {
        "variant_a": {
            "overall": overall_a,
            "skills": {k: v.to_dict() for k, v in scores_a.items()},
            "rubric_name": rubric.name if hasattr(rubric, "name") else "Current",
            "rubric_version": rubric.version,
        },
        "extracted_skills": [
            {
                "skill_name": m.get("skill_name"),
                "evidence_sentences": m.get("evidence_sentences", []),
                "quality": m.get("quality"),
            }
            for m in mapped
        ],
    }

    if scores_b:
        score_list_b = [s.final_score for s in scores_b.values()]
        overall_b = int(sum(score_list_b) / len(score_list_b)) if score_list_b else 0
        result["variant_b"] = {
            "overall": overall_b,
            "skills": {k: v.to_dict() for k, v in scores_b.items()},
            "rubric_name": "Variant B",
        }
        result["delta"] = overall_b - overall_a

    return result


def _mock_extract(text: str) -> List[Dict[str, Any]]:
    """Simple mock extraction for preview when no LLM is available.
    Uses keyword heuristics to guess skills from text.
    """
    text_lower = text.lower()
    tech_keywords = {
        "python": {"name": "Python", "evidence": []},
        "react": {"name": "React", "evidence": []},
        "javascript": {"name": "JavaScript", "evidence": []},
        "docker": {"name": "Docker", "evidence": []},
        "kubernetes": {"name": "Kubernetes", "evidence": []},
        "sql": {"name": "SQL", "evidence": []},
        "api": {"name": "API Design", "evidence": []},
        "node": {"name": "Node.js", "evidence": []},
        "aws": {"name": "AWS", "evidence": []},
        "git": {"name": "Git", "evidence": []},
    }

    extracted = []
    for kw, info in tech_keywords.items():
        if kw in text_lower:
            sentences = [
                s.strip()
                for s in text.replace("\\n", ".").split(".")
                if kw in s.lower()
            ]
            extracted.append(
                {
                    "skill_name": info["name"],
                    "evidence_sentences": sentences[:2],
                }
            )
    return extracted[:5]


# =============================================================================
# PHASE 2 — A/B TESTING
# =============================================================================


@router.post("/ab-test/create")
def create_ab_test(
    body: dict,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Create a new A/B experiment for a job."""
    job_id = body.get("job_id")
    if not job_id:
        raise HTTPException(status_code=400, detail="job_id is required")
    _ensure_job_access(recruiter, job_id, db)

    variant_a = body.get("variant_a")
    variant_b = body.get("variant_b")
    if not variant_a or not variant_b:
        raise HTTPException(
            status_code=400, detail="Both variant_a and variant_b are required"
        )

    experiment = ABTestExperiment(
        job_id=job_id,
        created_by=recruiter.id,
        name=body.get("name", "Untitled Experiment"),
        description=body.get("description", ""),
        variant_a_json=variant_a,
        variant_b_json=variant_b,
        traffic_split=body.get("traffic_split", 50),
        min_sample_size=body.get("min_sample_size", 50),
    )
    db.add(experiment)
    db.commit()
    db.refresh(experiment)
    return {"id": experiment.id, "name": experiment.name, "status": experiment.status}


@router.post("/ab-test/{experiment_id}/start")
def start_ab_test(
    experiment_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Start an A/B experiment."""
    exp = (
        db.query(ABTestExperiment).filter(ABTestExperiment.id == experiment_id).first()
    )
    if not exp:
        raise HTTPException(status_code=404, detail="Experiment not found")
    _ensure_job_access(recruiter, exp.job_id, db)
    exp.status = "running"
    exp.started_at = datetime.utcnow()
    db.commit()
    return {"id": exp.id, "status": "running", "started_at": exp.started_at.isoformat()}


@router.post("/ab-test/{experiment_id}/stop")
def stop_ab_test(
    experiment_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Stop an A/B experiment."""
    exp = (
        db.query(ABTestExperiment).filter(ABTestExperiment.id == experiment_id).first()
    )
    if not exp:
        raise HTTPException(status_code=404, detail="Experiment not found")
    _ensure_job_access(recruiter, exp.job_id, db)
    exp.status = "completed"
    exp.ended_at = datetime.utcnow()
    db.commit()
    return {"id": exp.id, "status": "completed"}


@router.post("/ab-test/assign-user")
def assign_ab_user(
    body: dict,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Assign a user/candidate to an experiment variant.
    Uses consistent hashing for deterministic assignment when user_id is stable.
    """
    experiment_id = body.get("experiment_id")
    user_id = body.get("user_id")
    candidate_id = body.get("candidate_id")

    if not experiment_id:
        raise HTTPException(status_code=400, detail="experiment_id is required")
    if not user_id and not candidate_id:
        raise HTTPException(status_code=400, detail="user_id or candidate_id required")

    exp = (
        db.query(ABTestExperiment).filter(ABTestExperiment.id == experiment_id).first()
    )
    if not exp:
        raise HTTPException(status_code=404, detail="Experiment not found")
    _ensure_job_access(recruiter, exp.job_id, db)
    if exp.status != "running":
        raise HTTPException(status_code=400, detail="Experiment is not running")

    identity = str(user_id or candidate_id)
    hash_val = int(hashlib.md5(f"{experiment_id}:{identity}".encode()).hexdigest(), 16)
    variant = "a" if (hash_val % 100) < exp.traffic_split else "b"

    existing = (
        (
            db.query(ABTestAssignment)
            .filter(
                ABTestAssignment.experiment_id == experiment_id,
                ABTestAssignment.user_id == user_id,
            )
            .first()
        )
        if user_id
        else None
    )

    if not existing and candidate_id:
        existing = (
            db.query(ABTestAssignment)
            .filter(
                ABTestAssignment.experiment_id == experiment_id,
                ABTestAssignment.candidate_id == candidate_id,
            )
            .first()
        )

    if existing:
        return {
            "experiment_id": experiment_id,
            "variant": existing.variant,
            "existing": True,
        }

    assignment = ABTestAssignment(
        experiment_id=experiment_id,
        user_id=user_id,
        candidate_id=candidate_id,
        variant=variant,
    )
    db.add(assignment)
    exp.current_sample_size = (
        db.query(ABTestAssignment)
        .filter(ABTestAssignment.experiment_id == experiment_id)
        .count()
    )
    db.commit()

    return {"experiment_id": experiment_id, "variant": variant, "existing": False}


@router.get("/ab-test/{experiment_id}/results")
def get_ab_results(
    experiment_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Get A/B test results with statistics."""
    exp = (
        db.query(ABTestExperiment).filter(ABTestExperiment.id == experiment_id).first()
    )
    if not exp:
        raise HTTPException(status_code=404, detail="Experiment not found")
    _ensure_job_access(recruiter, exp.job_id, db)

    results = (
        db.query(ScoringVariantResult)
        .filter(ScoringVariantResult.experiment_id == experiment_id)
        .all()
    )

    scores_a = [r.variant_a_score for r in results]
    scores_b = [r.variant_b_score for r in results]
    deltas = [r.score_delta for r in results if r.score_delta is not None]
    preferences = [r.recruiter_preference for r in results if r.recruiter_preference]

    def _stats(scores: List[int]) -> Dict:
        if not scores:
            return {"mean": 0, "median": 0, "std": 0, "min": 0, "max": 0, "n": 0}
        n = len(scores)
        mean = sum(scores) / n
        sorted_s = sorted(scores)
        median = (
            sorted_s[n // 2] if n % 2 else (sorted_s[n // 2 - 1] + sorted_s[n // 2]) / 2
        )
        variance = sum((s - mean) ** 2 for s in scores) / n
        return {
            "mean": round(mean, 1),
            "median": round(median, 1),
            "std": round(variance**0.5, 1),
            "min": min(scores),
            "max": max(scores),
            "n": n,
        }

    pref_a = preferences.count("a") if preferences else 0
    pref_b = preferences.count("b") if preferences else 0

    return {
        "experiment_id": experiment_id,
        "experiment_name": exp.name,
        "status": exp.status,
        "sample_size": len(results),
        "min_sample_size": exp.min_sample_size,
        "variant_a": {**_stats(scores_a), "config": exp.variant_a_json},
        "variant_b": {**_stats(scores_b), "config": exp.variant_b_json},
        "delta_stats": _stats(deltas) if deltas else None,
        "recruiter_preferences": {
            "variant_a": pref_a,
            "variant_b": pref_b,
            "total": pref_a + pref_b,
        },
    }


@router.get("/ab-test/list")
def list_ab_tests(
    job_id: Optional[int] = Query(None),
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """List A/B tests for a recruiter or job."""
    q = db.query(ABTestExperiment).filter(ABTestExperiment.created_by == recruiter.id)
    if job_id:
        q = q.filter(ABTestExperiment.job_id == job_id)
    tests = q.order_by(ABTestExperiment.created_at.desc()).all()
    return [
        {
            "id": t.id,
            "job_id": t.job_id,
            "name": t.name,
            "status": t.status,
            "traffic_split": t.traffic_split,
            "sample_size": t.current_sample_size,
            "min_sample_size": t.min_sample_size,
            "created_at": t.created_at.isoformat(),
        }
        for t in tests
    ]


@router.post("/ab-test/{experiment_id}/record-result")
def record_ab_result(
    experiment_id: int,
    body: dict,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Record a scoring result pair from an A/B experiment."""
    exp = (
        db.query(ABTestExperiment).filter(ABTestExperiment.id == experiment_id).first()
    )
    if not exp:
        raise HTTPException(status_code=404, detail="Experiment not found")
    _ensure_job_access(recruiter, exp.job_id, db)
    result = ScoringVariantResult(
        experiment_id=experiment_id,
        candidate_id=body["candidate_id"],
        variant_a_score=body["variant_a_score"],
        variant_b_score=body["variant_b_score"],
        variant_a_json=body.get("variant_a_json", {}),
        variant_b_json=body.get("variant_b_json", {}),
        score_delta=body.get("variant_b_score", 0) - body.get("variant_a_score", 0),
        recruiter_preference=body.get("recruiter_preference"),
    )
    db.add(result)
    exp.current_sample_size = (
        db.query(ScoringVariantResult)
        .filter(ScoringVariantResult.experiment_id == experiment_id)
        .count()
        + 1
    )
    db.commit()
    return {"id": result.id, "score_delta": result.score_delta}


@router.get("/coverage/{application_id}")
def get_rubric_coverage(
    application_id: int,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Return rubric coverage metrics for a completed interview."""
    get_application_for_recruiter(application_id, recruiter, db)
    summary = (
        db.query(EvaluationResult)
        .join(
            EvaluationSession,
            EvaluationResult.evaluation_session_id == EvaluationSession.id,
        )
        .filter(EvaluationSession.application_id == application_id)
        .first()
    )
    if not summary:
        raise HTTPException(
            status_code=404, detail="No rubric summary found for this interview"
        )

    breakdown = summary.score_breakdown or {}
    if isinstance(breakdown, str):
        breakdown = json.loads(breakdown) if breakdown else {}
    bd = breakdown if isinstance(breakdown, dict) else {}

    categories = bd.get("category_scores", []) or []
    skills_assessed = 0
    skills_total = 0
    for cat in categories:
        skills_assessed += cat.get("skills_scored", 0) if isinstance(cat, dict) else 0
        skills_total += cat.get("skills_total", 0) if isinstance(cat, dict) else 0

    return {
        "interview_id": application_id,
        "overall_score": summary.final_score,
        "skills_assessed": skills_assessed,
        "skills_total": skills_total,
        "coverage_pct": round((skills_assessed / skills_total) * 100)
        if skills_total > 0
        else 0,
        "gaps": bd.get("gaps", []) or [],
        "num_answers_scored": bd.get("num_answers_scored", 0),
    }


# =============================================================================
# PHASE 3 — SKILL DEFINITION CRUD
# =============================================================================


@router.get("/skills")
def list_skills(
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
):
    """List canonical skill definitions with pagination."""
    total = db.query(func.count(SkillDefinitionDB.id)).scalar()
    skills = (
        db.query(SkillDefinitionDB)
        .order_by(SkillDefinitionDB.name)
        .offset(skip)
        .limit(limit)
        .all()
    )
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "skills": [
            {
                "id": s.id,
                "name": s.name,
                "description": s.description,
                "expected_proficiency": s.expected_proficiency,
                "weight": s.weight,
                "keywords": s.keywords,
                "levels": s.levels,
                "is_required": s.is_required,
                "created_at": s.created_at.isoformat() if s.created_at else None,
            }
            for s in skills
        ],
    }


@router.post("/skills")
def create_skill(
    body: dict,
    request: Request,
    recruiter: User = Depends(require_recruiter),
    db: Session = Depends(get_db),
):
    """Create a new canonical skill definition."""
    name = body.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Skill name is required")
    existing = (
        db.query(SkillDefinitionDB).filter(SkillDefinitionDB.name == name).first()
    )
    if existing:
        raise HTTPException(status_code=409, detail=f"Skill '{name}' already exists")
    from uuid import uuid4

    skill = SkillDefinitionDB(
        id=str(uuid4()),
        name=name,
        description=body.get("description", ""),
        expected_proficiency=body.get("expected_proficiency", "mid"),
        weight=body.get("weight", 1.0),
        keywords=body.get("keywords"),
        levels=body.get("levels"),
        is_required=body.get("is_required", False),
    )
    db.add(skill)
    db.commit()
    db.refresh(skill)

    audit = AuditLog(
        user_id=recruiter.id,
        company_id=getattr(recruiter, "_company_id", None),
        action="skill_create",
        target_id=str(skill.id),
        details="Created skill '%s'" % name,
        ip_address=request.client.host if request.client else None,
    )
    db.add(audit)
    db.commit()

    return {
        "id": skill.id,
        "name": skill.name,
        "description": skill.description,
        "expected_proficiency": skill.expected_proficiency,
        "weight": skill.weight,
        "keywords": skill.keywords,
        "levels": skill.levels,
        "is_required": skill.is_required,
    }


@router.put("/skills/{skill_id}")
def update_skill(
    skill_id: str,
    body: dict,
    request: Request,
    recruiter: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Update a canonical skill definition."""
    skill = db.query(SkillDefinitionDB).filter(SkillDefinitionDB.id == skill_id).first()
    if not skill:
        raise HTTPException(status_code=404, detail="Skill not found")
    old_name = skill.name
    if "name" in body:
        skill.name = body["name"]
    if "description" in body:
        skill.description = body["description"]
    if "expected_proficiency" in body:
        skill.expected_proficiency = body["expected_proficiency"]
    if "weight" in body:
        skill.weight = body["weight"]
    if "keywords" in body:
        skill.keywords = body["keywords"]
    if "levels" in body:
        skill.levels = body["levels"]
    if "is_required" in body:
        skill.is_required = body["is_required"]
    db.commit()

    audit = AuditLog(
        user_id=recruiter.id,
        company_id=getattr(recruiter, "_company_id", None),
        action="skill_update",
        target_id=str(skill_id),
        details="Updated skill '%s'" % (skill.name or old_name),
        ip_address=request.client.host if request.client else None,
    )
    db.add(audit)
    db.commit()

    return {
        "id": skill.id,
        "name": skill.name,
        "description": skill.description,
        "expected_proficiency": skill.expected_proficiency,
        "weight": skill.weight,
        "keywords": skill.keywords,
        "levels": skill.levels,
        "is_required": skill.is_required,
    }


@router.delete("/skills/{skill_id}")
def delete_skill(
    skill_id: str,
    request: Request,
    recruiter: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete a canonical skill definition."""
    skill = db.query(SkillDefinitionDB).filter(SkillDefinitionDB.id == skill_id).first()
    if not skill:
        raise HTTPException(status_code=404, detail="Skill not found")
    skill_name = skill.name
    db.delete(skill)
    db.commit()

    audit = AuditLog(
        user_id=recruiter.id,
        company_id=getattr(recruiter, "_company_id", None),
        action="skill_delete",
        target_id=str(skill_id),
        details="Deleted skill '%s'" % skill_name,
        ip_address=request.client.host if request.client else None,
    )
    db.add(audit)
    db.commit()

    return {"success": True}
